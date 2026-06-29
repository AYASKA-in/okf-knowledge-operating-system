import io
from pathlib import Path

import openpyxl

from app.ingestion.models import ParsedDocument, Section
from app.ingestion.parsers.base import DocumentParser


class XlsxParser(DocumentParser):
    def parse(self, data: bytes, filename: str = "") -> ParsedDocument:
        title = Path(filename).stem if filename else "Untitled Spreadsheet"
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

        sections: list[Section] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                vals = [str(c) if c is not None else "" for c in row]
                rows.append(" | ".join(vals))
            text = "\n".join(rows).strip()
            if text:
                sections.append(Section(title=sheet_name, text=text))

        wb.close()

        return ParsedDocument(
            title=title,
            sections=sections,
            source_type="xlsx",
            metadata={"filename": filename, "sheets": len(wb.sheetnames)},
        )
